from requests_html import HTMLSession
from openpyxl import Workbook
import openpyxl
from langdetect import detect

#Gathering of World News articles from the New York Times website
session = HTMLSession()
url = 'https://www.nytimes.com/section/world'
r = session.get(url)
r.html.render(sleep=1, scrolldown=0)
articles = r.html.find('article')
newslist = []

#Putting the data into a dictionary array
for item in articles:
    try:
        newsitem = item.find('h2', first=True)
        newsdesc = item.find('p', first=True)
        newsarticle = {
            'title' : newsitem.text,
            'link' : newsitem.absolute_links,
            'desc' : newsdesc.text
        }
        newslist.append(newsarticle)
    except:
        pass

#Writing of gathered details onto an Excel spreadsheet.
path = "C:\\Users\PC\\Desktop\\Projects"
try:
    f = open(path + '\\WorldNews.xlsx')
    f.close()
except FileNotFoundError:
    wb = Workbook()

wb = openpyxl.load_workbook(path + '\\WorldNews.xlsx')
wb.save(path + "\\WorldNews.xlsx")
st = wb.active

header = st.cell(row = 1, column = 1)
header.value = "World News from the New York Times website"

count = 2
#Inserting of data into cells.
for x in newslist:
    #If the title is in a language that is not English, it will be skipped.
    if(detect(x['title']) != 'en'):
        continue
    input = st.cell(row = count, column = 1)
    input.value = x["title"]
    input = st.cell(row = count, column = 2)
    input.value = str(x["link"])
    input = st.cell(row = count, column = 3)
    input.value = x["desc"]
    count += 1

wb.save(path + "\\WorldNews.xlsx")